import os
import tempfile
from openpyxl import Workbook, load_workbook
from zipfile import BadZipFile
from openpyxl.utils.exceptions import InvalidFileException

try:
    # 在少數環境沒有 GUI 時，避免匯入就報錯
    from tkinter import messagebox as _msgbox
except Exception:
    _msgbox = None

FILENAME = "meeting_schedule.xlsx"

REQUIRED_SHEETS = {
    "Schedule": ["流水號", "預約日期", "時段ID", "會議ID", "預約人ID", "使用目的", "取消狀態"],
    "MeetingRooms": ["會議ID", "名稱", "帳號", "密碼", "連結", "用途", "停用狀態", "可外借狀態"],
    "TimeSlots": ["時間ID", "時間區段", "停用狀態", "備註"],
    "TempLock": ["UserID", "Date", "SlotID", "RoomID", "Status", "Timestamp"],
    "FixedBooking": ["BookingID", "Weekday", "SlotID", "RoomID", "UserID", "Purpose", "Canceled"]
}

import os
import re
from tkinter import messagebox

def guess_excel_opener_from_lockfile(xlsx_path: str) -> str | None:
    """
    嘗試從 Excel 的鎖定檔 `~$` 讀取並猜測開啟檔案的使用者
    """
    folder, name = os.path.dirname(xlsx_path), os.path.basename(xlsx_path)
    lock_name = "~$" + name
    lock_path = os.path.join(folder, lock_name)

    if not os.path.exists(lock_path):
        return None

    data = None
    for enc in ("utf-16-le", "utf-8", "mbcs", "latin-1"):
        try:
            with open(lock_path, "rb") as f:
                raw = f.read()
            text = raw.decode(enc, errors="ignore")
            data = text
            if data and data.strip():
                break
        except Exception:
            continue

    if not data:
        return None

    candidates = re.findall(r"[^\x00-\x1f\x7f]{3,}", data)
    if not candidates:
        return None

    # 優先包含中文或空白/點號的字串（比較像人名或 AD 帳號顯示名）
    def score(s: str) -> int:
        has_cjk = any('\u4e00' <= ch <= '\u9fff' for ch in s)
        has_space = " " in s
        return (2 if has_cjk else 0) + (1 if has_space else 0) + len(s) // 8

    candidates.sort(key=score, reverse=True)
    guess = candidates[0].strip()
    return guess if 3 <= len(guess) <= 64 else None


def is_excel_file_locked(filepath):
    """
    嘗試用 append 模式開檔；若被 Excel 佔用會觸發 PermissionError。
    若檔案被鎖住，會讀取鎖定檔並顯示使用者名稱。
    """
    try:
        with open(filepath, "a"):
            return False  # 沒有被鎖
    except PermissionError:
        # 如果檔案被鎖定，讀取鎖定檔來猜測是誰開啟的
        opener = guess_excel_opener_from_lockfile(filepath)
        if opener:
            messagebox.showerror("檔案鎖定", f"Excel 檔案正在被 {opener} 使用中，請先關閉該檔案再試。")
        else:
            messagebox.showerror("檔案鎖定", "Excel 檔案正在被使用中，請先關閉再重試。")
        return True  # 被鎖住


# ===== 初學者重點：更安全的存檔 =====
def is_excel_file_locked(filepath: str) -> bool:
    """
    嘗試用 append 模式開檔；若被 Excel 佔用會觸發 PermissionError。
    只用來提示「請先關閉 Excel 再重試」。
    """
    try:
        with open(filepath, "a"):
            return False
    except PermissionError:
        return True

def safe_save(wb, filename):
    """
    初學者註解：
    1) 嘗試直接把活頁簿 wb 儲存到 filename。
    2) 如果檔案被 Excel 開著（Windows 常見），會拋出 PermissionError，
       我們回傳 False，讓呼叫端顯示「請先關閉 Excel 再試」的提醒視窗。
    3) 其他非預期錯誤也回傳 False，交由呼叫端決定要不要另外處理與提示。
    """
    try:
        wb.save(filename)  # 嘗試直接覆寫檔案
        return True
    except PermissionError:
        # Excel 正在使用該檔案（檔案被鎖），無法寫入
        return False
    except Exception:
        # 其他錯誤，保持安靜回傳 False，由呼叫端決定是否提示
        return False

def init_excel_file(filename=FILENAME):
    """檢查並初始化 Excel 檔案與所有必要工作表（僅改用 safe_save，流程不變）"""
    if not os.path.exists(filename):
        wb = Workbook()
    else:
        try:
            wb = load_workbook(filename)
        except (BadZipFile, InvalidFileException):
            print("⚠️ Excel 檔案錯誤，刪除並重建")
            os.remove(filename)
            wb = Workbook()

    for sheet_name, headers in REQUIRED_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            # ✅ 加入範例資料（只在新建 sheet 時做一次）
            if sheet_name == "MeetingRooms":
                ws.append([
                    "Zoom99", "第一會議室", "ID:9999", "pW:88888",
                    "https://meet.test/room1", "範例", "FALSE", "TRUE"
                ])
            elif sheet_name == "TimeSlots":
                ws.append([1, "09:00–10:00", "FALSE", "範例"])

    if "Schedule" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Schedule")

    # 僅將 wb.save 改為 safe_save；失敗時維持你原本的對話框與退出行為
    if not safe_save(wb, filename):
        print(f"無法儲存 Excel 檔案，可能已被其他程式鎖定：{filename}")
        if _msgbox:
            _msgbox.showerror("存檔失敗", f"無法儲存 Excel 檔案。\n請先關閉檔案再重試：\n\n{filename}")
        exit()

def get_workbook(filename=FILENAME):
    return load_workbook(filename)
