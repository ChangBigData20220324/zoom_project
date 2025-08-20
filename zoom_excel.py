#zoom_excel
from weekly_overview import PageWeeklyOverview
from openpyxl import Workbook, load_workbook
import os
import uuid
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime,timedelta
from excel_manager import init_excel_file, FILENAME, safe_save

# ✅ 鎖定狀態與時間（秒）
LOCK_STATUS = "LOCKING"
LOCK_EXPIRY_SECONDS = 180 # 3分鐘

# 全域變數儲存跨畫面資料
app_state = {
    "selected_date": None,
    "selected_slots": [],
    "selected_room": None,
    "user_id": "",
    "purpose": "",
    "has_locked": False,
    "lock_token": str(uuid.uuid4())  # ✅ 新增唯一識別碼
}
def is_excel_file_locked(filepath):
    try:
        with open(filepath, "a"):
            return False  # 沒有被鎖
    except PermissionError:
        return True  # 被其他程式鎖住（例如 Excel）


# 載入可用時段
def load_time_slots(filename=FILENAME):
    wb = load_workbook(filename)
    ws = wb["TimeSlots"]
    time_slots = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        slot_id, time_range, disabled, *_ = row
        if str(disabled).strip().upper() != "TRUE":
            time_slots[int(slot_id)] = time_range
    return time_slots

def get_temp_locked_rooms(date, slot_ids, filename="meeting_schedule.xlsx"):
    from openpyxl import load_workbook

    LOCK_EXPIRY_SECONDS = 180
    try:
        wb = load_workbook(filename)
    except:
        return set()

    if "TempLock" not in wb.sheetnames:
        return set()

    ws = wb["TempLock"]
    now = datetime.now()
    locked_rooms = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            lock_date, slot_id, room_id, status, timestamp_str = row[1], row[2], row[3], row[4], row[5]
            if not (lock_date and slot_id and room_id and timestamp_str):
                continue
            lock_time = datetime.strptime(timestamp_str, "%Y/%m/%d %H:%M:%S")
            if lock_date == date and int(slot_id) in slot_ids:
                elapsed = (now - lock_time).total_seconds()
                if elapsed < LOCK_EXPIRY_SECONDS:
                    locked_rooms.add(room_id)
        except:
            continue

    return locked_rooms

def get_available_rooms(date, slot_ids):
    wb = load_workbook(FILENAME)
    ws_schedule = wb["Schedule"]
    ws_rooms = wb["MeetingRooms"]

    # 已被正式預約的房間
    booked_rooms = set()
    for row in ws_schedule.iter_rows(min_row=2, values_only=True):
        record_date, existing_slots, meeting_id, canceled = row[1], row[2], row[3], row[6]
        if record_date == date and not canceled:
            booked_slots = set(map(int, existing_slots.split(',')))
            if any(slot in booked_slots for slot in slot_ids):
                booked_rooms.add(meeting_id)

    # 取得正在鎖定中的房間
    locked_rooms = get_temp_locked_rooms(date, slot_ids)

    # 加入固定預約佔用的房間
    fixed_locked_rooms = set()
    weekday_index = datetime.strptime(date, "%Y/%m/%d").weekday()
    weekday_str = ["週一", "週二", "週三", "週四", "週五"][weekday_index]

    if "FixedBooking" in wb.sheetnames:
        ws_fixed = wb["FixedBooking"]
        for row in ws_fixed.iter_rows(min_row=2, values_only=True):
            if len(row) >= 7 and row[6] is True:
                continue  # ✅ 加上這行，跳過已取消的固定預約
            _, wday, sid, rid, *_ = row
            if wday == weekday_str and rid and sid in slot_ids:
                fixed_locked_rooms.add(rid)

    # ✅ 只根據「可外借狀態」為 TRUE 的房間
    available_rooms = []
    for row in ws_rooms.iter_rows(min_row=2, values_only=True):
        room_id, name, acc, pwd, link, usage, closed, allow_external = row

        if str(allow_external).strip().upper() != "TRUE":
            continue  # ✅ 若不可外借，直接跳過

        if room_id in booked_rooms or room_id in fixed_locked_rooms:
            continue  # 已被預約

        is_locked = room_id in locked_rooms
        available_rooms.append((room_id, name, usage, is_locked))

    return available_rooms


# 檢查某會議室在指定日期與時段是否已被預約，與excel做出比對
def is_conflict(date, slot_ids, room_id):
    wb = load_workbook(FILENAME)
    ws = wb["Schedule"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record_date, slots, booked_room, canceled = row[1], row[2], row[3], row[6]
        if record_date == date and booked_room == room_id and not canceled:
            booked_slots = set(map(int, slots.split(',')))
            if any(slot in booked_slots for slot in slot_ids):
                return True
    # ✅ 加上固定預約的衝突檢查（插在 return False 前面）
    if is_fixed_booked_on_date(date, slot_ids, room_id):
        return True
    return False
def is_fixed_booked_on_date(date: str, slot_ids: list, room_id: str) -> bool:
    # 將日期轉為對應的星期幾（0=週一）
    weekday_index = datetime.strptime(date, "%Y/%m/%d").weekday()
    weekday_str = ["週一", "週二", "週三", "週四", "週五"][weekday_index]

    wb = load_workbook(FILENAME)
    ws = wb["FixedBooking"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 7 and row[6] is True:
            continue  # ✅ 跳過已取消的固定預約
        _, wday, sid, rid, *_ = row
        if wday == weekday_str and rid == room_id and sid in slot_ids:
            return True

    return False

def find_schedule_conflicts_by_weekday(weekday_str, slot_ids, room_id):
    weekday_map = {"週一": 0, "週二": 1, "週三": 2, "週四": 3, "週五": 4}
    target_weekday = weekday_map.get(weekday_str)
    today = datetime.today().date()
    conflicts = []

    wb = load_workbook(FILENAME)
    ws = wb["Schedule"]
    ws_slot = wb["TimeSlots"]

    # 建立 slot_id → 時間區間 的對照表
    slot_time_map = {}
    for row in ws_slot.iter_rows(min_row=2, values_only=True):
        sid, time_str, disabled, *_ = row
        if sid and time_str:
            slot_time_map[int(sid)] = time_str

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 7:
            continue  # 資料異常，跳過

        booking_id, date_str, slots, booked_room, uid, purpose, canceled = row

        try:
            record_date = datetime.strptime(date_str, "%Y/%m/%d").date()
        except:
            continue

        if record_date < today or canceled or booked_room != room_id:
            continue

        if record_date.weekday() != target_weekday:
            continue

        booked_slots = set(map(int, slots.split(',')))
        for sid in slot_ids:
            if sid in booked_slots:
                slot_time = slot_time_map.get(sid, "時間未知")
                conflicts.append({
                    "date": date_str,
                    "slot": sid,
                    "slot_time": slot_time,
                    "room": room_id,
                    "user": uid,
                    "purpose": purpose,
                    "source": "Schedule"
                })

    return conflicts

# === 新增：固定預約送出前，比對 TempLock（將日期→週幾） ===
from datetime import datetime

_ZH_WEEKDAYS = ["週一", "週二", "週三", "週四", "週五", "週六", "週日"]  # 與 UI 字面一致

def _to_zh_weekday(dt):
    """
    初學者註解：
    - datetime.weekday()：週一=0、週日=6
    - 轉成中文字，方便和固定預約選單直接比對
    """
    return _ZH_WEEKDAYS[dt.weekday()]

def _parse_date_cell(cell):
    """
    初學者註解：
    - TempLock 的日期欄可能是字串或日期，統一轉成 datetime（只取日期）
    """
    if cell is None:
        return None
    if isinstance(cell, datetime):
        return datetime(cell.year, cell.month, cell.day)
    # 可能是 date
    try:
        from datetime import date as _date
        if isinstance(cell, _date):
            return datetime(cell.year, cell.month, cell.day)
    except Exception:
        pass
    if isinstance(cell, str):
        for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                d = datetime.strptime(cell.strip(), fmt)
                return datetime(d.year, d.month, d.day)
            except ValueError:
                continue
    return None

def _parse_ts_cell(cell):
    """
    初學者註解：
    - 解析 TempLock 的 Timestamp，判斷是否過期（TTL）
    """
    if cell is None:
        return None
    if isinstance(cell, datetime):
        return cell
    if isinstance(cell, str):
        for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(cell.strip(), fmt)
            except ValueError:
                continue
    return None

def _normalize_slot_ids(slot_ids):
    """
    初學者註解：
    - 讓時段 ID 變成 set[int]，方便交集比對
    """
    if slot_ids is None:
        return set()
    if isinstance(slot_ids, (list, tuple, set)):
        try:
            return {int(x) for x in slot_ids}
        except Exception:
            return set()
    try:
        return {int(slot_ids)}
    except Exception:
        return set()

def has_templock_conflict_for_fixed(weekday_str, slot_ids, room_id, filename=FILENAME):
    """
    功能（給固定預約送出前用）：
    - 從 Excel 讀取『TempLock』，把每筆鎖的「Date」轉為「週幾」，
      與固定預約的〈room_id, slot_ids, weekday_str〉做比對。
    - 只讀檢查，不寫入任何資料。
    回傳：
    - (True, message) 代表命中（有人正在一般預約）；(False, "") 代表無命中。
    """
    # 正規化參數
    try:
        target_room_id_int = int(room_id)
    except Exception:
        target_room_id_int = None  # 若無法轉 int，後面改用字串比對
    target_room_id_str = str(room_id).strip()
    target_slots = _normalize_slot_ids(slot_ids)
    if not target_slots:
        return False, ""

    # 開檔／無表直接視為無衝突
    try:
        wb = load_workbook(filename)
    except Exception:
        return False, ""
    if "TempLock" not in wb.sheetnames:
        return False, ""

    ws = wb["TempLock"]
    now = datetime.now()

    # 預期欄位順序：["UserID/Token", "Date", "SlotID", "RoomID", "Status", "Timestamp"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
        token, lock_date_cell, slot_id_cell, lock_room_cell, status_cell, ts_cell = row[:6]

        # 必要欄位要有
        if lock_date_cell is None or slot_id_cell is None or lock_room_cell is None or ts_cell is None:
            continue

        # TTL：過期就略過
        ts = _parse_ts_cell(ts_cell)
        if ts is None:
            continue
        if (now - ts).total_seconds() >= LOCK_EXPIRY_SECONDS:
            continue

        # 狀態欄位（若有值）必須等於 LOCK_STATUS
        if status_cell and str(status_cell).strip().upper() != str(LOCK_STATUS).strip().upper():
            continue

        # 會議室比對（容錯：同時嘗試 int 與字串）
        try:
            lock_room_int = int(lock_room_cell)
        except Exception:
            lock_room_int = None
        lock_room_str = str(lock_room_cell).strip()

        same_room = (
            (target_room_id_int is not None and lock_room_int == target_room_id_int)
            or (lock_room_str == target_room_id_str)
        )
        if not same_room:
            continue

        # 時段比對
        try:
            lock_slot_id = int(slot_id_cell)
        except Exception:
            continue
        if lock_slot_id not in target_slots:
            continue

        # 日期→週幾
        lock_dt = _parse_date_cell(lock_date_cell)
        if lock_dt is None:
            continue
        lock_wd = _to_zh_weekday(lock_dt)

        # 週幾命中 → 擋
        if str(lock_wd) == str(weekday_str):
            return True, f"此時段有人正在一般預約中（{lock_wd} / 房 {lock_room_str} / 時段 {lock_slot_id}）。"

    return False, ""


# 寫入預約
def add_booking():
    wb = load_workbook(FILENAME)
    ws = wb["Schedule"]
    new_id = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], int):
            new_id = max(new_id, row[0] + 1)
    slot_str = ",".join(map(str, app_state["selected_slots"]))
    ws.append([new_id, app_state["selected_date"], slot_str, app_state["selected_room"],
           app_state["user_id"], app_state["purpose"], False])
    # 初學者註解：改用安全存檔；若被 Excel 鎖住或寫入失敗，給人看得懂的訊息
    if not safe_save(wb, FILENAME):
        messagebox.showerror("錯誤", "儲存 Excel 檔案失敗，請先關閉 Excel 或檢查檔案權限。")
        return False  # ✅ 新增
    return True  # ✅ 新增
# 檢查某固定預約是否會與既有的 Schedule 衝突
def is_fixed_booking_conflict(weekday_str, slot_id, room_id, weeks_ahead=4):
    weekday_map = {
        "週一": 0,
        "週二": 1,
        "週三": 2,
        "週四": 3,
        "週五": 4
    }

    if weekday_str not in weekday_map:
        return False

    weekday_index = weekday_map[weekday_str]
    today = datetime.today()
    this_weekday = today.weekday()
    days_until_target = (weekday_index - this_weekday) % 7
    next_target_date = today + timedelta(days=days_until_target)

    wb = load_workbook(FILENAME)
    ws = wb["Schedule"]

# 回傳與 FixedBooking 衝突的清單（根據 星期、時段、會議室）
def find_fixed_conflicts(weekday_str, slot_ids, room_id):
    wb = load_workbook(FILENAME)
    ws_fixed = wb["FixedBooking"]
    ws_slot = wb["TimeSlots"]

    # 建立 slot_id → 時間區間 的對照表
    slot_time_map = {}
    for row in ws_slot.iter_rows(min_row=2, values_only=True):
        sid, time_str, disabled, *_ = row
        if sid and time_str:
            slot_time_map[int(sid)] = time_str

    conflicts = []

    for row in ws_fixed.iter_rows(min_row=2, values_only=True):
        if len(row) >= 7 and row[6] is True:
            continue  # ✅ 跳過已取消的固定預約

        bid, wday, sid, rid, uid, purpose = row[:6]

        if wday == weekday_str and sid in slot_ids and rid == room_id:
            conflicts.append({
                "source": "FixedBooking",
                "booking_id": bid,
                "weekday": wday,
                "slot": sid,
                "slot_time": slot_time_map.get(sid, "時間未知"),
                "room": rid,
                "user": uid,
                "purpose": purpose
            })

    return conflicts

# 回傳與 Schedule 衝突的清單（檢查未來 N 週）
def find_schedule_conflicts(weekday_str, slot_ids, room_id, weeks_ahead=4):
    weekday_map = {"週一": 0, "週二": 1, "週三": 2, "週四": 3, "週五": 4}
    if weekday_str not in weekday_map:
        return []

    target_weekday = weekday_map[weekday_str]
    today = datetime.today()
    days_until_next = (target_weekday - today.weekday() + 7) % 7
    base_date = today + timedelta(days=days_until_next)

    wb = load_workbook(FILENAME)
    ws = wb["Schedule"]
    conflicts = []

    for i in range(weeks_ahead):
        target_date = (base_date + timedelta(days=i * 7)).strftime("%Y/%m/%d")
        for row in ws.iter_rows(min_row=2, values_only=True):
            booking_id, record_date, slots, booked_room, uid, purpose, canceled = row[:7]
            if record_date == target_date and booked_room == room_id and not canceled:
                booked_slots = set(map(int, slots.split(',')))
                if any(slot in booked_slots for slot in slot_ids):
                    intersect_slots = booked_slots.intersection(slot_ids)
                    for sid in intersect_slots:
                        conflicts.append({
                            "source": "Schedule",
                            "booking_id": booking_id,
                            "date": record_date,
                            "slot": sid,
                            "room": booked_room,
                            "user": uid,
                            "purpose": purpose
                        })
    return conflicts


# 主應用程式
class MeetingApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("會議室預約系統")

        # 設定為 1/4 螢幕大小，適用常見桌面解析度
        self.geometry("960x540")
        self.resizable(False, False)
        self.configure(bg="white")

        # 重要：讓各 page 撐滿視窗
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (PageDateInput, PageTimeSelect, PageRoomSelect, PageConfirm, PageFinish, PageWeeklyOverview,PageCancelBooking,PageUserBookingList,PageCancelSuccess,PageFixedBooking,PageFixedCancelBooking,PageRoomInfo):
            page_name = F.__name__
            frame = F(parent=self, controller=self)
            self.frames[page_name] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("PageDateInput")


    def show_frame(self, page_name):
        frame = self.frames[page_name]
        frame.tkraise()
        if hasattr(frame, "refresh"):
            frame.refresh()
# ====== 新增頁面：會議室基本資訊（只讀清單 + 單顆〔複製〕）======
class PageRoomInfo(tk.Frame):
    """
    同一張「表格」內呈現：第 0 列是表頭，後面每列是資料。
    6 欄：會議名稱｜用途｜帳號｜密碼｜連結｜操作（複製）
    特別處理：帳號/密碼若在 Excel 以數值儲存，轉成「非科學記號」的字串，盡量貼近 Excel 實際儲存值。
    """
    N_COLS = 6
    MIN_SPECS = [140, 180, 120, 110, 260, 90]  # 每欄最小寬度
    COL_WEIGHTS = [2, 3, 2, 2, 5, 1]            # 欄位比例（連結較寬）

    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller

        # 中段清單撐滿空間
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # ====== 頁面標題列與返回按鈕 ======
        header = tk.Frame(self, bg="white")
        header.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
        header.grid_columnconfigure(0, weight=1)
        tk.Label(header, text="會議室基本資訊（只讀）",
                 font=("Arial", 14, "bold"), bg="white", fg="#111827").grid(row=0, column=0, sticky="w")
        tk.Button(header, text="返回首頁",
                  command=lambda: controller.show_frame("PageDateInput"),
                  font=("Arial", 11), bg="#e5e7eb", fg="#111827", relief="flat"
        ).grid(row=0, column=1, sticky="e")

        # ====== 可滾動表格（表頭與資料同一張表）======
        outer = tk.Frame(self, bg="white")
        outer.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))

        canvas = tk.Canvas(outer, bg="white", highlightthickness=0)
        vbar = tk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        self.rows_frame = tk.Frame(canvas, bg="white")  # 這個 Frame 內就是整張表（表頭 + 資料）

        self._rows_window = canvas.create_window((0, 0), window=self.rows_frame, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(self._rows_window, width=e.width))
        self.rows_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.configure(yscrollcommand=vbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        vbar.pack(side="right", fill="y")

        # 滑鼠滾輪
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # 收集需要動態設定 wraplength 的儲存格（每欄一個 list）
        self._cells_by_col = {0: [], 1: [], 2: [], 3: [], 4: []}

        # rows_frame 寬度變化時，依比例與最小寬度設定每欄文字的換行寬度
        def _on_table_resize(e):
            total_min = sum(self.MIN_SPECS)
            total_wt = sum(self.COL_WEIGHTS)
            leftover = max(0, e.width - total_min)
            for col, labels in self._cells_by_col.items():
                col_width = self.MIN_SPECS[col] + leftover * (self.COL_WEIGHTS[col] / total_wt)
                wrap = max(1, int(col_width) - 12)  # 扣掉左右 padding
                for lbl in labels:
                    lbl.configure(wraplength=wrap)
        self.rows_frame.bind("<Configure>", _on_table_resize)

    # --------- 內部工具：欄配置與儲存格元件 ---------

    def _config_table_columns(self):
        """設定整張表（rows_frame）的欄配置：最小寬度 + 權重"""
        for i in range(self.N_COLS):
            self.rows_frame.grid_columnconfigure(i,
                                                 minsize=self.MIN_SPECS[i],
                                                 weight=self.COL_WEIGHTS[i])

    def _head_cell(self, col, text):
        """建立表頭儲存格"""
        lbl = tk.Label(self.rows_frame, text=text, bg="#f9fafb", fg="#111827",
                       font=("Arial", 10, "bold"),
                       bd=1, relief="solid", padx=6, pady=8,
                       anchor="w", justify="left")
        lbl.grid(row=0, column=col, sticky="nsew")

    def _data_cell(self, row, col, text, fg="#111827"):
        """建立資料儲存格（一般文字）"""
        lbl = tk.Label(self.rows_frame, text=text, bg="white", fg=fg,
                       bd=1, relief="solid", padx=6, pady=6,
                       anchor="w", justify="left", wraplength=1)
        lbl.grid(row=row, column=col, sticky="nsew")
        if col in self._cells_by_col:
            self._cells_by_col[col].append(lbl)

    # --------- 內部工具：把 Excel 的帳號/密碼轉成「非科學記號」字串 ---------

    def _excel_num_as_text(self, v):
        """
        將 Excel 以數值儲存的帳號/密碼轉為字串顯示：
        - 若是 int：直接 str(v)
        - 若是 float：
            * 近似整數 → 使用不帶科學記號的整數格式（避免 1.23E+16）
            * 其他情況 → 以 Decimal 轉成十進位字串，去除多餘的 0
        - 其他型別（字串/None）→ 正常處理
        說明：Excel 對超過 15 位的數字會以雙精度儲存，精度已在 Excel 層流失；此處僅避免顯示成科學記號。
        """
        from decimal import Decimal
        if v is None:
            return ""
        if isinstance(v, int):
            return str(v)
        if isinstance(v, float):
            # 若非常接近整數，當整數顯示（避免 3.0、1.23E+16）
            if abs(v - round(v)) < 1e-9:
                return "{:.0f}".format(v)
            # 其他浮點數：用 Decimal 轉成一般十進位字串
            d = Decimal(str(v))
            s = format(d, "f")
            return s.rstrip("0").rstrip(".") if "." in s else s
        # 若原本就是字串（例如 'pW:88888'），保留原貌但去前後空白
        return str(v).strip()

    def refresh(self):
        # 清表：移除所有儲存格與行，再重建表頭與資料
        for w in self.rows_frame.winfo_children():
            w.destroy()
        for k in self._cells_by_col:
            self._cells_by_col[k].clear()

        self._config_table_columns()

        # ====== 表頭（row=0）======
        self._head_cell(0, "會議名稱")
        self._head_cell(1, "用途")
        self._head_cell(2, "帳號")
        self._head_cell(3, "密碼")
        self._head_cell(4, "連結")
        self._head_cell(5, "操作")

        # ====== 讀取資料並建立每列（row 從 1 開始）======
        from openpyxl import load_workbook
        wb = load_workbook(FILENAME)
        if "MeetingRooms" not in wb.sheetnames:
            tk.Label(self.rows_frame, text="找不到 MeetingRooms 工作表",
                     bg="white", fg="#6b7280").grid(row=1, column=0, columnspan=self.N_COLS, sticky="w", padx=6, pady=6)
            return
        ws = wb["MeetingRooms"]

        def _s(v):  # 一般文字欄位：None -> "" 並去空白
            return "" if v is None else str(v).strip()

        row_grid = 1
        # values_only=False 取得 Cell 物件；但此需求只需值即可
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Excel 欄位順序：0:ID, 1:名稱, 2:帳號, 3:密碼, 4:連結, 5:用途, 6:停用, 7:可外借
            name_raw, account_raw, password_raw, url_raw, purpose_raw, disabled = row[1], row[2], row[3], row[4], row[5], row[6]

            disabled_flag = (disabled is True) or (str(disabled).strip().upper() == "TRUE")
            if disabled_flag:
                continue

            # 轉成要顯示的文字（帳號/密碼特別處理成「非科學記號」）
            name = _s(name_raw)
            purpose = _s(purpose_raw)
            account = self._excel_num_as_text(account_raw)
            password = self._excel_num_as_text(password_raw)
            url = _s(url_raw)

            # 擺上同一個 rows_frame（同一張表）
            self._data_cell(row_grid, 0, name)
            self._data_cell(row_grid, 1, purpose)
            self._data_cell(row_grid, 2, account)
            self._data_cell(row_grid, 3, password)
            self._data_cell(row_grid, 4, url, fg="#2563eb")

            # 操作欄：有邊框的儲存格 + 內嵌按鈕
            op_cell = tk.Frame(self.rows_frame, bg="white", bd=1, relief="solid")
            op_cell.grid(row=row_grid, column=5, sticky="nsew")
            tk.Button(op_cell, text="複製",
                      command=lambda a=account, p=password, u=url: self.copy_triplet(a, p, u),
                      font=("Arial", 10), bg="#f3f4f6", fg="#111827", relief="flat"
            ).pack(fill="x", padx=6, pady=6)

            row_grid += 1

        if row_grid == 1:
            tk.Label(self.rows_frame, text="目前沒有啟用中的會議室",
                     bg="white", fg="#6b7280").grid(row=1, column=0, columnspan=self.N_COLS, sticky="w", padx=6, pady=6)

    def copy_triplet(self, account, password, url):
        """把帳號/密碼/連結三項依固定格式寫入剪貼簿。確保與畫面顯示一致（非科學記號）。"""
        def _s(v):
            return "" if v is None else str(v).strip()
        # account/password 已是顯示用字串，這裡不再轉型；只對 url 做一般去空白
        text = f"帳號: {account}\n密碼: {password}\n連結: {_s(url)}"
        self.clipboard_clear()
        self.clipboard_append(text)


#98 預約記錄查詢
class PageUserBookingList(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.user_id = ""
        self.records = []
        self.vars = []

        # 頁面標題
        self.lbl_title = tk.Label(
            self,
            text="使用者預約紀錄",
            font=("Arial", 18, "bold"),
            bg="white",
            fg="#111827"
        )
        self.lbl_title.pack(anchor="w", padx=40, pady=(30, 10))


        # 清單區
        # 外層 Frame（保留 padding）
        outer_frame = tk.Frame(self, bg="white")
        outer_frame.pack(fill="both", expand=True, padx=40, pady=10)

        # Canvas + Scrollbar
        canvas = tk.Canvas(outer_frame, bg="white", highlightthickness=0)
        scrollbar = tk.Scrollbar(outer_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, bg="white")

        # 將內容 frame 放進 canvas
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 排版
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 讓舊程式中引用 self.list_frame 的地方不需改動
        self.list_frame = self.scrollable_frame
        # 滑鼠滾輪支援：Windows
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # 滑鼠滾輪支援：Linux (可選)
        canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))


        # 操作按鈕
        btn_frame = tk.Frame(self, bg="white")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="刪除選取", command=self.delete_selected, font=("Arial", 12),
                  bg="#ef4444", fg="white", relief="flat").pack(side="left", padx=10)

        tk.Button(btn_frame, text="返回", command=lambda: controller.show_frame("PageCancelBooking"),
                  font=("Arial", 12), bg="#e5e7eb", fg="#111827", relief="flat").pack(side="left", padx=10)

    def load_user_records(self, user_id):
        from datetime import datetime, timedelta

        self.user_id = user_id
        self.lbl_title.config(text=f"使用者：{user_id}，預約紀錄")
        self.records = []
        self.vars = []

        for widget in self.list_frame.winfo_children():
            widget.destroy()

        slot_mapping = self.load_time_slot_mapping()
        today = datetime.today().date()

        wb = load_workbook(FILENAME)
        ws = wb["Schedule"]

        # === 一般預約 ===
        for row in ws.iter_rows(min_row=2, values_only=True):
            booking_id, date_str, slots, room_id, uid, purpose, canceled = row
            if uid == user_id and not canceled:
                try:
                    booking_date = datetime.strptime(str(date_str), "%Y/%m/%d").date()
                    if booking_date >= today:
                        slot_ids = list(map(int, slots.split(',')))
                        time_strings = [slot_mapping.get(sid, f"時段{sid}") for sid in slot_ids]
                        time_str = "，".join(time_strings)

                        # ✅ 每新增一筆預約記錄時，也新增一個對應的勾選變數
                        self.records.append(("[一般]", booking_id, booking_date, time_str, room_id, purpose))
                        var = tk.BooleanVar()
                        self.vars.append(var)
                except:
                    continue



        # ✅ 日期排序
        self.records.sort(key=lambda r: r[2])  # r[2] 是 datetime 物件

        # ✅ 顯示
        if not self.records:
            tk.Label(self.list_frame, text="查無任何即將到來的預約紀錄", font=("Arial", 12), bg="white", fg="gray")\
                .pack(pady=20)
            return

        

        if not self.records:
            tk.Label(self.list_frame, text="查無任何即將到來的預約紀錄", font=("Arial", 12), bg="white", fg="gray")\
                .pack(pady=20)
            return


                 # ✅ 注意：這裡不可以再縮排
        for i, record in enumerate(self.records):
            label_type, booking_id, dt, time_str, room_id, purpose = record
            date = dt.strftime("%Y/%m/%d")
            var = self.vars[i]

            row_frame = tk.Frame(self.list_frame, bg="white")
            row_frame.pack(fill="x", pady=6, anchor="w")

            cb = tk.Checkbutton(
                row_frame,
                variable=var,
                bg="#fee2e2",
                font=("Arial", 30),
                selectcolor="white",
                activebackground="white",
                relief="flat",
                padx=6,
                pady=4
            )
            cb.pack(side="left", padx=(0, 10), pady=2)

            right_frame = tk.Frame(row_frame, bg="white")
            right_frame.pack(side="left", fill="both", expand=True)

            label_type = "[一般]"
            bg_color = "#fef3c7"

            lbl_date = tk.Label(
                right_frame,
                text=f"{label_type} [{booking_id}] 日期：{date}",
                bg=bg_color,
                fg="#111827",
                font=("Arial", 11, "bold"),
                anchor="w",
                justify="left",
                padx=6,
                pady=4
            )
            lbl_date.pack(fill="x")

            text = f"時間：{time_str}\n會議室：{room_id}\n用途：{purpose}"
            lbl_info = tk.Label(
                right_frame,
                text=text,
                bg="white",
                anchor="w",
                justify="left",
                wraplength=720,
                font=("Arial", 11)
            )
            lbl_info.pack(fill="x", pady=(2, 0))

    def load_time_slot_mapping(self):
        # 讀取 TimeSlots 工作表，建立 {1: "09:00–10:00", 2: "..."} 對照表
        wb = load_workbook(FILENAME)
        ws = wb["TimeSlots"]
        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            slot_id, time_range, disabled, *_ = row  # 使用 *_ 來忽略後面的多餘欄位
            mapping[int(slot_id)] = time_range
        return mapping

    def delete_selected(self):
        selected_ids = [
            self.records[i][1] for i, var in enumerate(self.vars) if var.get()
        ]
        if not selected_ids:
            messagebox.showwarning("提醒", "請勾選要刪除的預約")
            return

        # 初學者註解：
        # 1) 先檢查 Excel 是否被開啟（被鎖住），鎖住時就不繼續往下做，避免做了也存不進去。
        if is_excel_file_locked(FILENAME):
            messagebox.showerror("錯誤", "Excel 檔案正在被使用中，請先關閉再試。")
            return

        try:
            wb = load_workbook(FILENAME)
            ws = wb["Schedule"]
            updated = False

            # ✅ 一般預約取消（將第 7 欄 canceled 設為 True）
            for row in ws.iter_rows(min_row=2):
                if row[0].value in selected_ids:
                    row[6].value = True
                    updated = True

            if not updated:
                messagebox.showinfo("提示", "找不到對應的預約資料，可能已被取消或不存在。")
                return

            # 初學者註解：
            # 2) 嘗試存檔，safe_save() 回傳 False 代表「沒有成功寫入」（例如檔案被鎖）
            ok = safe_save(wb, FILENAME)
            if not ok:
                messagebox.showerror("錯誤", "儲存 Excel 檔案失敗，請先關閉 Excel 或檢查檔案權限。")
                return

            # 只有在確認寫入成功後，才顯示成功頁面
            self.controller.show_frame("PageCancelSuccess")

        except Exception as e:
            messagebox.showerror("錯誤", f"讀取或寫入 Excel 檔案失敗。\n錯誤原因：{e}")



#99 取消頁面
class PageCancelBooking(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        main_frame = tk.Frame(self, bg="white")
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.grid_rowconfigure(1, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)

        # 頁面標題
        tk.Label(
            main_frame,
            text="【一般取消】請輸入預約人 ID",
            font=("Arial", 16, "bold"),
            bg="white",
            fg="red"
        ).grid(row=0, column=0, sticky="w", padx=40, pady=(30, 10))

        # 表單主體區
        form_frame = tk.Frame(main_frame, bg="white")
        form_frame.grid(row=1, column=0, sticky="nsew", padx=40, pady=20)
        form_frame.grid_rowconfigure((0, 1, 2), weight=1)
        form_frame.grid_columnconfigure(0, weight=1)

        # Step 1：提示文字
        tk.Label(
            form_frame,
            text="請輸入您的預約人 ID（如：U001）：",
            font=("Arial", 12),
            bg="white"
        ).grid(row=0, column=0, sticky="w", padx=10, pady=(5, 0))

        # 紅框輸入欄
        border_frame = tk.Frame(form_frame, bg="red")
        border_frame.grid(row=1, column=0, sticky="ew", padx=40, pady=(5, 20))

        self.entry_userid = tk.Entry(
            border_frame,
            font=("Arial", 20),
            justify="center",
            relief="flat",
            bg="white",
            fg="black"
        )
        self.entry_userid.pack(fill="both", expand=True, ipadx=10, ipady=10, padx=2, pady=2)

        # Step 2：查詢按鈕
        btn_frame = tk.Frame(form_frame, bg="white")
        btn_frame.grid(row=2, column=0)

        # 查詢按鈕外紅框
        btn_border = tk.Frame(btn_frame, highlightbackground="red", highlightthickness=2, bg="white")
        btn_border.pack(pady=(10, 10))

        tk.Button(
            btn_border,
            text="查詢預約紀錄",
            font=("Arial", 12),
            width=18,
            bg="#3b82f6",
            fg="white",
            relief="flat",
            command=self.query_user_booking  # ⚠️ 後續可實作查詢功能
        ).pack(padx=2, pady=2)

        # 返回首頁
        tk.Button(
            main_frame,
            text="返回首頁",
            command=lambda: controller.show_frame("PageDateInput"),
            font=("Arial", 12),
            bg="#e5e7eb",
            fg="#111827",
            relief="flat"
        ).grid(row=2, column=0, pady=(0, 30))

    def query_user_booking(self):
        user_id = self.entry_userid.get().strip()
        if not user_id:
            messagebox.showwarning("提醒", "請輸入預約人 ID")
            return

        self.controller.frames["PageUserBookingList"].load_user_records(user_id)
        self.controller.show_frame("PageUserBookingList")
class PageFixedBooking(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.time_slots = load_time_slots()
        self.rooms = self.load_rooms()

        tk.Label(self, text="固定預約管理", font=("Arial", 20, "bold"), bg="white").pack(pady=(30, 10))

        # ========== 主容器 ========== #
        main_container = tk.Frame(self, bg="white")
        main_container.pack(fill="both", expand=True, padx=40, pady=(0, 0))
        main_container.grid_rowconfigure(0, weight=1)
        main_container.grid_rowconfigure(1, weight=0)
        main_container.grid_columnconfigure(0, weight=1)

        # ========== 上方左右分區 ========== #
        content_frame = tk.Frame(main_container, bg="white")
        content_frame.grid(row=0, column=0, sticky="nsew")

        content_frame.grid_columnconfigure(0, weight=1)
        content_frame.grid_columnconfigure(1, weight=1)

        left_frame = tk.Frame(content_frame, bg="white")
        right_frame = tk.Frame(content_frame, bg="white")
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 30))
        right_frame.grid(row=0, column=1, sticky="nsew")

        # ===== 左側欄位 =====
        tk.Label(left_frame, text="預約人 ID：", font=("Arial", 12), bg="white").pack(anchor="w", pady=(0, 5))
        self.entry_user = tk.Entry(left_frame, font=("Arial", 12), width=28)
        self.entry_user.pack(pady=(0, 15))

        tk.Label(left_frame, text="使用目的：", font=("Arial", 12), bg="white").pack(anchor="w", pady=(0, 5))
        self.text_purpose = tk.Text(left_frame, font=("Arial", 12), width=28, height=10)
        self.text_purpose.pack()

        # ===== 右側欄位 =====
        # 星期選單（中文 + 請選擇）
        tk.Label(right_frame, text="星期幾：", font=("Arial", 12), bg="white").pack(anchor="w", pady=(0, 5))
        self.weekday_var = tk.StringVar()
        weekdays = ["請選擇", "週一", "週二", "週三", "週四", "週五"]
        self.weekday_var.set("請選擇")
        weekday_menu = tk.OptionMenu(right_frame, self.weekday_var, *weekdays)
        weekday_menu.config(font=("Arial", 12), width=24)
        weekday_menu.pack(pady=(0, 15))

        # 會議室選單（請選擇）
        tk.Label(right_frame, text="會議室：", font=("Arial", 12), bg="white").pack(anchor="w", pady=(0, 5))
        self.room_var = tk.StringVar()
        room_ids = [r[0] for r in self.rooms]
        room_options = ["請選擇"] + room_ids
        self.room_var.set("請選擇")
        room_menu = tk.OptionMenu(right_frame, self.room_var, *room_options)
        room_menu.config(font=("Arial", 12), width=24)
        room_menu.pack(pady=(0, 15))

        # 時段選擇（兩欄排）
        tk.Label(right_frame, text="選擇時段：", font=("Arial", 12), bg="white").pack(anchor="w", pady=(0, 5))
        slot_container = tk.Frame(right_frame, bg="white")
        slot_container.pack()

        self.slot_vars = {}
        left_col = tk.Frame(slot_container, bg="white")
        right_col = tk.Frame(slot_container, bg="white")
        left_col.grid(row=0, column=0, padx=(0, 30), sticky="n")
        right_col.grid(row=0, column=1, sticky="n")

        sorted_slots = sorted(self.time_slots.items())
        half = (len(sorted_slots) + 1) // 2
        for i, (sid, label) in enumerate(sorted_slots):
            var = tk.BooleanVar()
            target_col = left_col if i < half else right_col
            cb = tk.Checkbutton(
                target_col,
                text=f"{sid}: {label}",
                variable=var,
                bg="white",
                font=("Arial", 10),
                anchor="w"
            )
            cb.pack(anchor="w")
            self.slot_vars[sid] = var

        # ========== 下方按鈕區 ========== #
        btn_frame = tk.Frame(main_container, bg="white")
        btn_frame.grid(row=1, column=0, pady=(20, 20))

        tk.Button(
            btn_frame,
            text="儲存固定預約",
            command=self.save_fixed_booking,
            font=("Arial", 12),
            bg="#10b981",
            fg="white",
            relief="flat",
            width=18
        ).pack(side="left", padx=20)

        tk.Button(
            btn_frame,
            text="返回首頁",
            command=lambda: controller.show_frame("PageDateInput"),
            font=("Arial", 12),
            bg="#e5e7eb",
            fg="#111827",
            relief="flat",
            width=14
        ).pack(side="left", padx=20)

    def load_rooms(self):
        wb = load_workbook(FILENAME)
        ws = wb["MeetingRooms"]
        return [
            row for row in ws.iter_rows(min_row=2, values_only=True)
            if str(row[6]).strip().upper() != "TRUE"  # ✅ 排除「停用」的會議室
        ]


    def save_fixed_booking(self):
        # 取得表單欄位內容
        weekday = self.weekday_var.get().strip()
        room_id = self.room_var.get().strip()
        user_id = self.entry_user.get().strip()
        purpose = self.text_purpose.get("1.0", "end").strip()
        selected_slots = [sid for sid, var in self.slot_vars.items() if var.get()]

        # ===== 防呆驗證 =====
        if weekday == "請選擇" or not weekday:
            messagebox.showwarning("欄位不完整", "請選擇星期幾")
            return
        if room_id == "請選擇" or not room_id:
            messagebox.showwarning("欄位不完整", "請選擇會議室")
            return
        if not selected_slots:
            messagebox.showwarning("欄位不完整", "請至少選擇一個時段")
            return
        if not user_id:
            messagebox.showwarning("欄位不完整", "請輸入預約人 ID")
            return
        if not purpose:
            messagebox.showwarning("欄位不完整", "請輸入使用目的")
            return

        # === 新增：TempLock（一般預約進行中）比對，一旦命中就阻擋本次送出 ===
        hit, _msg = has_templock_conflict_for_fixed(
            weekday_str=weekday,
            slot_ids=selected_slots,
            room_id=room_id,          # 可為字串；函式內會做容錯
            filename=FILENAME
        )
        if hit:
            messagebox.showerror("預約中", "此時段有人正在預約中，請稍候再送出。")
            return

        # ===== 衝突檢查（保留你原本的邏輯） =====
        fixed_conflicts = find_fixed_conflicts(weekday, selected_slots, room_id)
        schedule_conflicts = find_schedule_conflicts_by_weekday(weekday, selected_slots, room_id)

        if fixed_conflicts or schedule_conflicts:
            conflict_msgs = []
            for c in fixed_conflicts:
                conflict_msgs.append(
                    f"[固定] 星期 {c['weekday']}，時段 {c['slot']}, {c['slot_time']}，會議室 {c['room']}，預約人 {c['user']}（用途：{c['purpose']}）"
                )
            for c in schedule_conflicts:
                conflict_msgs.append(
                    f"[正式] 日期 {c['date']}，時段 {c['slot']}, {c['slot_time']}，會議室 {c['room']}，預約人 {c['user']}（用途：{c['purpose']}）"
                )
            messagebox.showerror(
                "衝突提醒",
                "以下預約已存在，請重新選擇時段或會議室：\n\n" + "\n".join(conflict_msgs)
            )
            return

        # ===== 寫入固定預約（原本流程） =====
        wb = load_workbook(FILENAME)
        ws = wb["FixedBooking"]
        max_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if isinstance(row[0], int):
                max_id = max(max_id, row[0])
        for sid in selected_slots:
            ws.append([max_id + 1, weekday, sid, room_id, user_id, purpose, False])
            max_id += 1
        # 初學者註解：固定預約 → 安全存檔
        if not safe_save(wb, FILENAME):
            messagebox.showerror("錯誤", "固定預約儲存失敗，請先關閉 Excel 或檢查檔案權限。")
            return

        messagebox.showinfo("成功", "固定預約已儲存")
        self.controller.show_frame("PageDateInput")

# 頁面一：輸入日期
class PageDateInput(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        main_frame = tk.Frame(self, bg="white")
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.grid_rowconfigure(1, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)

        # 功能選單區塊（灰底、有邊框、有標題）
        menu_frame = tk.Frame(main_frame, bg="#f3f4f6", highlightbackground="#d1d5db", highlightthickness=1)
        menu_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)

        # 標題：「功能選單」
        tk.Label(
            menu_frame,
            text="功能選單",
            font=("Arial", 10, "bold"),
            bg="#f3f4f6",
            fg="#111827"
        ).pack(anchor="w", padx=10, pady=(6, 4))

        # 內層按鈕框（平均分散按鈕）
        buttons_frame = tk.Frame(menu_frame, bg="#f3f4f6")
        buttons_frame.pack(fill="x", padx=10, pady=(0, 10))

        # 每個欄位都分到等寬空間
        for i in range(4):
            buttons_frame.grid_columnconfigure(i, weight=1)

        btn_cfg = {
            "bg": "white",
            "relief": "flat",
            "font": ("Arial", 10, "underline"),
            "cursor": "hand2",
            "padx": 10,
            "pady": 6
        }

        tk.Button(buttons_frame, text="固定預約管理", fg="#047857",
                command=lambda: controller.show_frame("PageFixedBooking"), **btn_cfg).grid(row=0, column=0, padx=4, sticky="ew")

        tk.Button(buttons_frame, text="本週預約狀況", fg="#2563eb",
                command=lambda: controller.show_frame("PageWeeklyOverview"), **btn_cfg).grid(row=0, column=1, padx=4, sticky="ew")

        tk.Button(buttons_frame, text="一般取消", fg="#dc2626",
                command=lambda: controller.show_frame("PageCancelBooking"), **btn_cfg).grid(row=0, column=2, padx=4, sticky="ew")

        tk.Button(buttons_frame, text="固定取消", fg="#dc2626",
                command=lambda: controller.show_frame("PageFixedCancelBooking"), **btn_cfg).grid(row=0, column=3, padx=4, sticky="ew")
        # 新增第 5 顆：會議室基本資訊
        tk.Button(
            buttons_frame,
            text="會議室基本資訊",
            fg="#111827",
            command=lambda: controller.show_frame("PageRoomInfo"),
            **btn_cfg
        ).grid(row=0, column=4, padx=4, sticky="ew")
        # 區隔線（功能選單與下方表單分離）
        separator = tk.Frame(main_frame, height=1, bg="#d1d5db")
        separator.grid(row=1, column=0, sticky="ew", padx=10)



        # 表單主體區
        form_frame = tk.Frame(main_frame, bg="white")
        form_frame.grid(row=1, column=0, sticky="nsew", padx=40, pady=20)
        form_frame.grid_rowconfigure((0, 1, 2, 3, 4, 5), weight=1)
        form_frame.grid_columnconfigure(0, weight=1)

        # Step 1：輸入日期標題
        tk.Label(
            form_frame,
            text="【Step 1】請輸入預約日期",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="red"
        ).grid(row=0, column=0, sticky="w", padx=10, pady=(5, 0))

        # 範例文字
        tk.Label(
            form_frame,
            text="範例：（2025/05/06）",
            font=("Arial", 10),
            bg="white",
            fg="#777"
        ).grid(row=1, column=0, sticky="w", padx=10)

        # 紅框輸入欄
        border_frame = tk.Frame(form_frame, bg="red", highlightthickness=0)
        border_frame.grid(row=2, column=0, padx=40, pady=(5, 20), sticky="ew")

        self.date_entry = tk.Entry(
            border_frame,
            font=("Arial", 20),
            justify="center",
            relief="flat",
            bg="white",
            fg="black"
        )
        self.date_entry.pack(fill="both", expand=True, ipadx=10, ipady=10, padx=2, pady=2)

        # Step 2：操作引導
        tk.Label(
            form_frame,
            text="【Step 2】點選下方按鈕查詢或清空",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="red"
        ).grid(row=3, column=0, sticky="w", padx=10, pady=(10, 5))

        # 按鈕列使用 grid 安全置中
        btn_row = tk.Frame(form_frame, bg="white")
        btn_row.grid(row=4, column=0, pady=(5, 10))
        btn_row.grid_columnconfigure((0, 1, 2), weight=1)

        center_frame = tk.Frame(btn_row, bg="white")
        center_frame.grid(row=0, column=1)

        # 紅框包查詢按鈕
        btn_check_border = tk.Frame(
            center_frame,
            highlightbackground="red",
            highlightthickness=2,
            bg="white"
        )
        btn_check_border.pack(side="left", padx=(0, 10))

        tk.Button(
            btn_check_border,
            text="查詢可預約時段",
            command=self.next_page,
            font=("Arial", 12),
            width=18,
            bg="#3b82f6",
            fg="white",
            relief="flat"
        ).pack(padx=2, pady=2, ipadx=5)

        # 清空按鈕（不含紅框）
        tk.Button(
            center_frame,
            text="清空",
            command=self.clear_input,
            font=("Arial", 12),
            width=10,
            bg="#e5e7eb",
            fg="#111827",
            relief="flat"
        ).pack(side="left", padx=10, ipadx=5)

    def next_page(self):
        date_input = self.date_entry.get().strip()
        try:
            parsed_date = datetime.strptime(date_input, "%Y/%m/%d")
        except ValueError:
            messagebox.showerror("錯誤", "請輸入正確的日期格式（例如：2025/05/06）")
            return

        today = datetime.today()
        if parsed_date.date() < today.date():
            messagebox.showerror("錯誤", "只能預約今天或未來的日期")
            return
        if is_excel_file_locked(FILENAME):
            messagebox.showerror("錯誤", "Excel 檔案正在被使用中，請先關閉再試。")
            return
        app_state["selected_date"] = parsed_date.strftime("%Y/%m/%d")
        self.controller.show_frame("PageTimeSelect")




    def clear_input(self):
        self.date_entry.delete(0, tk.END)

# 頁面二：選擇可預約時段
class PageTimeSelect(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.vars = {}
        self.time_slots = load_time_slots()

        # 主容器
        main_frame = tk.Frame(self, bg="white")
        main_frame.pack(fill="both", expand=True, padx=40, pady=20)

        # Step 3 標題
        tk.Label(
            main_frame,
            text="【Step 3】請選擇可預約的時段",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="red"
        ).pack(anchor="w", pady=(0, 10))

        # 紅框包時段列表
        self.slot_border = tk.Frame(main_frame, bg="red")
        self.slot_border.pack(fill="x", pady=(0, 20))

        self.slot_frame = tk.Frame(self.slot_border, bg="white")
        self.slot_frame.pack(padx=2, pady=2, fill="x")

        # Step 4 標題
        tk.Label(
            main_frame,
            text="【Step 4】操作按鈕",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="red"
        ).pack(anchor="w", pady=(0, 10))

        # ✅ 三個按鈕同排（下一步有紅框）
        btn_row = tk.Frame(main_frame, bg="white")
        btn_row.pack(anchor="w")

        # 👉 紅框包住「下一步」
        btn_next_border = tk.Frame(btn_row, highlightbackground="red", highlightthickness=2, bg="white")
        btn_next_border.pack(side="left", padx=(0, 8))

        tk.Button(
            btn_next_border,
            text="下一步",
            command=self.next_page,
            font=("Arial", 12),
            width=12,
            bg="#3b82f6",
            fg="white",
            relief="flat"
        ).pack(padx=2, pady=2, ipadx=5)

        # 👉「清空」按鈕
        tk.Button(
            btn_row,
            text="清空",
            command=self.reset_slots,
            font=("Arial", 12),
            width=10,
            bg="#e5e7eb",
            fg="#111827",
            relief="flat"
        ).pack(side="left", padx=8, ipadx=5)

        # 👉「返回」按鈕
        tk.Button(
            btn_row,
            text="返回",
            command=lambda: controller.show_frame("PageDateInput"),
            font=("Arial", 12),
            width=10,
            bg="#e5e7eb",
            fg="#111827",
            relief="flat"
        ).pack(side="left", padx=8, ipadx=5)

    def refresh(self):
        cleanup_expired_locks()
        for widget in self.slot_frame.winfo_children():
            widget.destroy()
        self.vars.clear()

        date = app_state["selected_date"]
        available_slots = []

        for sid, time_str in self.time_slots.items():
            rooms = get_available_rooms(date, [sid])
            if rooms:
                available_slots.append((sid, time_str))

        if not available_slots:
            messagebox.showinfo("預約已滿", f"{date} 所有時段都已被預約，請選擇其他日期")
            self.controller.show_frame("PageDateInput")
            return
                # 改為兩欄垂直排列，左邊先塞 1、2、3，右邊接 4、5、6
        mid = (len(available_slots) + 1) // 2
        left_slots = available_slots[:mid]
        right_slots = available_slots[mid:]

        for row in range(max(len(left_slots), len(right_slots))):
            if row < len(left_slots):
                sid, time_str = left_slots[row]
                var = tk.IntVar()
                cb = tk.Checkbutton(
                    self.slot_frame,
                    text=f"時段 {sid}: {time_str}",
                    variable=var,
                    anchor="w",
                    bg="white",
                    font=("Arial", 12),
                    padx=10
                )
                cb.grid(row=row, column=0, sticky="w", padx=20, pady=4)
                self.vars[sid] = var

            if row < len(right_slots):
                sid, time_str = right_slots[row]
                var = tk.IntVar()
                cb = tk.Checkbutton(
                    self.slot_frame,
                    text=f"時段 {sid}: {time_str}",
                    variable=var,
                    anchor="w",
                    bg="white",
                    font=("Arial", 12),
                    padx=10
                )
                cb.grid(row=row, column=1, sticky="w", padx=20, pady=4)
                self.vars[sid] = var





        for sid in app_state["selected_slots"]:
            if sid in self.vars:
                self.vars[sid].set(1)
        
    def reset_slots(self):
        for var in self.vars.values():
            var.set(0)
        app_state["selected_slots"] = []

    def next_page(self):
        app_state["selected_slots"] = [sid for sid, var in self.vars.items() if var.get()]
        if not app_state["selected_slots"]:
            messagebox.showwarning("提醒", "請至少選擇一個時段")
            return
        if is_excel_file_locked(FILENAME):
            messagebox.showerror("錯誤", "Excel 檔案正在被使用中，請先關閉再試。")
            return
        self.controller.show_frame("PageRoomSelect")

# 頁面三：選擇會議室
class PageRoomSelect(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.rooms = []

        # 整體分成上下兩半：Step 5 + Step 6
        self.grid_rowconfigure(0, weight=1)  # Step 5：會議室選擇
        self.grid_rowconfigure(1, weight=1)  # Step 6：操作按鈕
        self.grid_columnconfigure(0, weight=1)

        # === Step 5：選擇會議室 ===
        upper_frame = tk.Frame(self, bg="white")
        upper_frame.grid(row=0, column=0, sticky="nsew", padx=40, pady=(20, 10))

        # 標題
        tk.Label(
            upper_frame,
            text="【Step 5】請選擇會議室",
            font=("Arial", 12, "bold"),
            fg="red",
            bg="white"
        ).pack(anchor="w", pady=(0, 10))

        # 外框留白容器
        outer_padding_frame = tk.Frame(upper_frame, bg="white")
        outer_padding_frame.pack(fill="both", expand=True, padx=25)  # ✅ 左右留白

        # 紅色邊框容器
        self.room_border = tk.Frame(outer_padding_frame, bg="red")
        self.room_border.pack(fill="both", expand=True)

        # 白底實際內容
        self.room_frame = tk.Frame(self.room_border, bg="white")
        self.room_frame.pack(padx=2, pady=2, fill="both", expand=True)

        # 會議室選擇清單
        self.listbox = tk.Listbox(
            self.room_frame,
            font=("Arial", 12),
            selectmode=tk.SINGLE,
            exportselection=False
        )
        self.listbox.pack(fill="both", expand=True, padx=10, pady=10)

        # === Step 6：操作按鈕 ===
        lower_frame = tk.Frame(self, bg="white")
        lower_frame.grid(row=1, column=0, sticky="nsew", padx=40, pady=(10, 20))

        # 標題
        tk.Label(
            lower_frame,
            text="【Step 6】操作按鈕",
            font=("Arial", 12, "bold"),
            fg="red",
            bg="white"
        ).pack(anchor="w", pady=(0, 10))

        # 按鈕列
        btn_row = tk.Frame(lower_frame, bg="white")
        btn_row.pack(anchor="w")

        # 👉 紅框包「下一步」
        btn_next_border = tk.Frame(btn_row, highlightbackground="red", highlightthickness=2, bg="white")
        btn_next_border.pack(side="left", padx=(0, 8))

        tk.Button(
            btn_next_border,
            text="下一步",
            command=self.next_page,
            font=("Arial", 12),
            width=12,
            bg="#3b82f6",
            fg="white",
            relief="flat"
        ).pack(padx=2, pady=2, ipadx=5)

        # 👉「返回」按鈕
        tk.Button(
            btn_row,
            text="返回",
            command=lambda: controller.show_frame("PageTimeSelect"),
            font=("Arial", 12),
            width=10,
            bg="#e5e7eb",
            fg="#111827",
            relief="flat"
        ).pack(side="left", padx=8, ipadx=5)
        # 位置：PageRoomSelect.__init__ 的下方按鈕列（就在『返回』旁邊）
        tk.Button(
            btn_row,
            text="重新整理",
            command=self.on_refresh,     # 初學者註解：按下就重跑 refresh()
            font=("Arial", 12),
            width=10,
            bg="#e5e7eb",
            fg="#111827",
            relief="flat"
        ).pack(side="left", padx=8, ipadx=5)

    # 同一類別中新增一個方法
    def on_refresh(self):
        """
        初學者註解：
        - 重新整理會做兩件事：
        1) cleanup_expired_locks()：清過期暫鎖，避免清單顯示過時
        2) 重新載入可用會議室：同步正式/固定預約的最新狀態
        """
        self.refresh()
        from tkinter import messagebox
        messagebox.showinfo("已更新", "已重新整理會議室清單")


    def refresh(self):
        # 清除過期暫時鎖定的資料
        cleanup_expired_locks()

        # 清除 Listbox 舊資料
        self.listbox.delete(0, tk.END)

        # 取得目前選擇的日期與時段
        date = app_state["selected_date"]
        slots = app_state["selected_slots"]
        
        # 查詢會議室清單（會包含 LOCKING 狀態）
        self.rooms = get_available_rooms(date, slots)

        # 若沒有任何可用的會議室
        if not self.rooms:
            self.listbox.insert(tk.END, "（無可用會議室）")
            self.listbox.config(state=tk.DISABLED)
            return
        else:
            self.listbox.config(state=tk.NORMAL)

        selected_index = None
        

        for idx, (room_id, name, usage, is_locked) in enumerate(self.rooms):
            display_text = f"{room_id} - {name}（{usage}）"
            self.listbox.insert(tk.END, display_text)

            if room_id == app_state.get("selected_room"):
                selected_index = idx

        # 回填原本選取的會議室（如果還在清單內）
        if selected_index is not None:
            self.listbox.selection_set(selected_index)
            self.listbox.activate(selected_index)
            self.listbox.see(selected_index)

    def next_page(self):
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning("提醒", "請選擇一間會議室")
            return

        index = selection[0]
        room_id = self.rooms[index][0]
        app_state["selected_room"] = room_id

        date = app_state["selected_date"]
        slot_ids = app_state["selected_slots"]
        my_token = app_state["lock_token"]

        # ✅ 檢查是否已被正式預約（避免 stale 資訊）
        if is_conflict(date, slot_ids, room_id):
            messagebox.showerror("預約失敗", "此會議室已被其他人正式預約，請重新選擇。")
            self.refresh()
            return

        # ✅ TempLock 再次確認是否有人預約中
        wb = load_workbook(FILENAME)
        ws = wb["TempLock"]
        now = datetime.now()
        for row in ws.iter_rows(min_row=2, values_only=True):
            token, lock_date, slot_id, locked_room, status, timestamp_str = row
            if lock_date == date and int(slot_id) in slot_ids and locked_room == room_id:
                try:
                    lock_time = datetime.strptime(timestamp_str, "%Y/%m/%d %H:%M:%S")
                    if (now - lock_time).total_seconds() < LOCK_EXPIRY_SECONDS and token != my_token:
                        messagebox.showerror("預約中", "此會議室正在被他人預約中，請等待約 3 分鐘後再試。")
                        return
                except:
                    continue

        # ✅ 若沒鎖 → 我來鎖起來
        lock_room(token=my_token, date=date, slot_ids=slot_ids, room_id=room_id)
        app_state["has_locked"] = True

        if is_excel_file_locked(FILENAME):
            messagebox.showerror("錯誤", "Excel 檔案正在被使用中，請先關閉再試。")
            return
        self.controller.show_frame("PageConfirm")



# 頁面四：輸入預約人資料
# 頁面四：輸入預約人資料
class PageConfirm(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller

        # 使用 grid 控制整體佈局
        self.grid_rowconfigure(0, weight=3)  # Step 7 (3/4)
        self.grid_rowconfigure(1, weight=1)  # Step 8 (1/4)
        self.grid_columnconfigure(0, weight=1)

        # Step 7：輸入欄位（占 3/4）
        top_frame = tk.Frame(self, bg="white")
        top_frame.grid(row=0, column=0, sticky="nsew", padx=40, pady=(20, 10))

        tk.Label(
            top_frame,
            text="【Step 7】請輸入預約資訊",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="red"
        ).pack(anchor="w", pady=(0, 10))

        input_border = tk.Frame(top_frame, bg="red")
        input_border.pack(fill="both", expand=True)

        input_frame = tk.Frame(input_border, bg="white")
        input_frame.pack(padx=2, pady=2, fill="both", expand=True)

        # 預約人 ID
        tk.Label(input_frame, text="預約人 ID：", font=("Arial", 12), bg="white").pack(anchor="w", pady=(10, 0), padx=10)
        self.entry_user = tk.Entry(input_frame, font=("Arial", 14))
        self.entry_user.pack(fill="x", padx=10, pady=(0, 15), ipady=10)

        # 使用目的：多行文字框
        tk.Label(input_frame, text="使用目的：", font=("Arial", 12), bg="white").pack(anchor="w", pady=(0, 0), padx=10)
        self.entry_purpose = tk.Text(input_frame, font=("Arial", 14), height=6, wrap="word")
        self.entry_purpose.pack(fill="both", expand=True, padx=10, pady=(0, 15))

        # Step 8：操作按鈕（占 1/4）
        bottom_frame = tk.Frame(self, bg="white")
        bottom_frame.grid(row=1, column=0, sticky="nsew", padx=40, pady=(10, 20))

        tk.Label(
            bottom_frame,
            text="【Step 8】操作按鈕",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="red"
        ).pack(anchor="w", pady=(0, 10))

        btn_row = tk.Frame(bottom_frame, bg="white")
        btn_row.pack(anchor="w")

        btn_next_border = tk.Frame(btn_row, highlightbackground="red", highlightthickness=2, bg="white")
        btn_next_border.pack(side="left", padx=(0, 8))

        tk.Button(
            btn_next_border,
            text="完成預約",
            command=self.finish,
            font=("Arial", 12),
            width=12,
            bg="#10b981",
            fg="white",
            relief="flat"
        ).pack(padx=2, pady=2, ipadx=5)

        tk.Button(
            btn_row,
            text="清空輸入",
            command=self.reset_fields,
            font=("Arial", 12),
            width=10,
            bg="#e5e7eb",
            fg="#111827",
            relief="flat"
        ).pack(side="left", padx=8, ipadx=5)

        tk.Button(
            btn_row,
            text="返回",
            command=self.cancel_and_back,
            font=("Arial", 12),
            width=10,
            bg="#e5e7eb",
            fg="#111827",
            relief="flat"
        ).pack(side="left", padx=8, ipadx=5)
    def cancel_and_back(self):
        # 如果有鎖定，就釋放
            if app_state["has_locked"]:
             release_token_locks(app_state["lock_token"])
            app_state["has_locked"] = False
            self.controller.show_frame("PageRoomSelect")


    def refresh(self):
        cleanup_expired_locks()
        self.entry_user.delete(0, tk.END)
        self.entry_user.insert(0, app_state.get("user_id", ""))

        self.entry_purpose.delete("1.0", tk.END)
        self.entry_purpose.insert("1.0", app_state.get("purpose", ""))
        self.start_timeout_timer()  # ✅ 加這行
    # 啟動或重設計時器
    def start_timeout_timer(self):
        if hasattr(self, "_timeout_id"):
            self.after_cancel(self._timeout_id)  # 若已有倒數 → 取消

        # 啟動新的倒數計時器（180秒 = 180,000 毫秒）
        self._timeout_id = self.after(180000, self.timeout_redirect)

    # 倒數結束 → 自動跳轉
    def timeout_redirect(self):
        if app_state["has_locked"]:
            release_token_locks(app_state["lock_token"])
            app_state["has_locked"] = False

        messagebox.showinfo("預約超時", "您停留太久，系統將自動返回首頁，請重新預約。")
        self.controller.show_frame("PageDateInput")

    

    def reset_fields(self):
        self.entry_user.delete(0, tk.END)
        self.entry_purpose.delete("1.0", tk.END)

    def finish(self):
        uid = self.entry_user.get().strip()
        purpose = self.entry_purpose.get("1.0", tk.END).strip()
        if not uid or not purpose:
            messagebox.showerror("錯誤", "請填寫所有欄位")
            return

        app_state["user_id"] = uid
        app_state["purpose"] = purpose

        booking_data = get_booking_data()
        date = booking_data["date"]
        slot_ids = booking_data["time_slots"]
        room_id = booking_data["room_id"]

        # 轉換時段 ID 為時間文字
        slot_map = load_time_slots()
        slot_texts = [slot_map.get(sid, f"時段 {sid}") for sid in slot_ids]
        slot_display = "，".join(slot_texts)

        confirm_msg = (
            f"【預約人 ID】：{uid}\n"
            f"【預約日期】：{date}\n"
            f"【預約時段】：{slot_display}\n"
            f"【會議室 ID】：{room_id}\n"
            f"【使用目的】：{purpose}\n\n"
            f"★請確認以上資訊正確，是否送出？"
        )

        if not messagebox.askyesno("確認預約", confirm_msg):
            return

        self.do_booking()
    def do_booking(self):
        booking_data = get_booking_data()
        date = booking_data["date"]
        slot_ids = booking_data["time_slots"]
        room_id = booking_data["room_id"]

        if is_conflict(date, slot_ids, room_id):
            messagebox.showerror(
                "預約失敗",
                f"您選擇的會議室（{room_id}）在指定時段已被其他人預約。\n請重新選擇。"
            )
            self.controller.show_frame("PageRoomSelect")
            return

        if not add_booking():   # ✅ 若寫入失敗，不執行後續
            return

        release_token_locks(app_state["lock_token"])
        app_state["has_locked"] = False
        app_state["lock_token"] = str(uuid.uuid4())

        self.entry_user.delete(0, tk.END)
        self.entry_purpose.delete("1.0", tk.END)

        for key in app_state:
            if isinstance(app_state[key], list):
                app_state[key] = []
            else:
                app_state[key] = ""

        self.controller.show_frame("PageFinish")



# 頁面五：完成畫面
class PageFinish(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller

        # 上區（標題區）：白色背景
        top_frame = tk.Frame(self, bg="white")
        top_frame.pack(fill="both", expand=True)

        # 下區（說明與按鈕）：淡灰色背景
        bottom_frame = tk.Frame(self, bg="#f3f4f6")
        bottom_frame.pack(fill="x")

        # ✅ 預約完成標題
        tk.Label(
            top_frame,
            text="✅ 預約完成！",
            font=("Arial", 36, "bold"),
            fg="#10b981",  # 綠色字
            bg="white"
        ).pack(pady=(100, 20))  # 上方多留白

        # 📌 提醒文字（保留資訊內容）
        tk.Label(
            bottom_frame,
            text="若需取消預約，請聯絡資訊部\n分機 : XXX 或來信 : XXXX",
            font=("Arial", 20),
            fg="#111827",  # 深灰字
            bg="#f3f4f6",
            justify="center"
        ).pack(pady=(30, 20))

        # 🔁 返回首頁按鈕
        tk.Button(
            bottom_frame,
            text="返回首頁",
            command=lambda: controller.show_frame("PageDateInput"),
            font=("Arial", 16),
            width=16,
            bg="#3b82f6",  # 藍底白字
            fg="white",
            relief="flat"
        ).pack(pady=(0, 40), ipadx=10, ipady=5)
# 將目前的 app_state 打包成乾淨 dict
def get_booking_data():
    return {
        "date": app_state["selected_date"],
        "time_slots": app_state["selected_slots"],
        "room_id": app_state["selected_room"],
        "user_id": app_state["user_id"],
        "purpose": app_state["purpose"]
    }
# ✅ 取消成功畫面（風格與預約成功一致）
class PageCancelSuccess(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller

        # 上半部（白底）
        top_frame = tk.Frame(self, bg="white")
        top_frame.pack(fill="both", expand=True)

        # 下半部（灰底）
        bottom_frame = tk.Frame(self, bg="#f3f4f6")
        bottom_frame.pack(fill="x")

        # ✅ 標題
        tk.Label(
            top_frame,
            text="✅ 預約已成功取消！",
            font=("Arial", 36, "bold"),
            fg="#ef4444",  # 紅色字
            bg="white"
        ).pack(pady=(100, 20))

        # 📌 提示說明
        tk.Label(
            bottom_frame,
            text="如需重新預約，請返回首頁",
            font=("Arial", 20),
            fg="#111827",
            bg="#f3f4f6",
            justify="center"
        ).pack(pady=(30, 20))

        # 🔁 返回首頁按鈕
        tk.Button(
            bottom_frame,
            text="返回首頁",
            command=lambda: controller.show_frame("PageDateInput"),
            font=("Arial", 16),
            width=16,
            bg="#3b82f6",
            fg="white",
            relief="flat"
        ).pack(pady=(0, 40), ipadx=10, ipady=5)
class PageFixedCancelBooking(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.records = []
        self.vars = []

        self.time_map = self.load_time_slot_mapping()

        tk.Label(self, text="固定預約取消", font=("Arial", 18, "bold"), bg="white",fg="red").pack(pady=(30, 10))

        # ==== 使用者輸入區 ====
        input_frame = tk.Frame(self, bg="white")
        input_frame.pack(pady=10)

        tk.Label(input_frame, text="請輸入預約人 ID：", font=("Arial", 12), bg="white").grid(row=0, column=0, sticky="w")
        self.entry_userid = tk.Entry(input_frame, font=("Arial", 14))
        self.entry_userid.grid(row=1, column=0, padx=10, pady=(0, 10))

        tk.Button(input_frame, text="查詢固定預約", font=("Arial", 12),
                  bg="#3b82f6", fg="white", relief="flat", command=self.search).grid(row=1, column=1, padx=10)

        # ==== 小標題顯示 ====
        self.lbl_result_title = tk.Label(self, text="", font=("Arial", 12, "bold"), bg="white", fg="#111827")
        self.lbl_result_title.pack(pady=(5, 5))

        # ==== 清單區（可滾動）====
        list_container = tk.Frame(self, bg="white")
        list_container.pack(fill="both", expand=True, padx=40, pady=10)

        canvas = tk.Canvas(list_container, bg="white", highlightthickness=0)
        scrollbar = tk.Scrollbar(list_container, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, bg="white")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.list_frame = self.scrollable_frame  # 為了相容原本的程式

        # ✅ 滾輪支援
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Linux 上
        canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

        # ==== 按鈕區 ====
        btn_frame = tk.Frame(self, bg="white")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="取消選取預約", font=("Arial", 12),
                  bg="#ef4444", fg="white", relief="flat", command=self.cancel_selected).pack(side="left", padx=10)

        tk.Button(btn_frame, text="返回首頁", font=("Arial", 12),
                  bg="#e5e7eb", fg="#111827", relief="flat",
                  command=lambda: controller.show_frame("PageDateInput")).pack(side="left", padx=10)

    def load_time_slot_mapping(self):
        wb = load_workbook(FILENAME)
        ws = wb["TimeSlots"]
        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid, time_str, disabled, *_ = row
            if sid and time_str:
                mapping[int(sid)] = time_str
        return mapping

    def search(self):
        user_id = self.entry_userid.get().strip()
        self.records = []
        self.vars = []

        for widget in self.list_frame.winfo_children():
            widget.destroy()

        self.lbl_result_title.config(text="")  # 清空標題

        if not user_id:
            messagebox.showwarning("提醒", "請輸入使用者 ID")
            return

        wb = load_workbook(FILENAME)
        if "FixedBooking" not in wb.sheetnames:
            messagebox.showinfo("無資料", "目前尚無任何固定預約")
            return

        ws = wb["FixedBooking"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 6 or row[4] != user_id:
                continue
            if len(row) >= 7 and row[6] is True:
                continue  # 已取消

            booking_id, wday, sid, rid, uid, purpose = row[:6]
            time_str = self.time_map.get(sid, f"時段 {sid}")
            text = f"ID: {booking_id}｜{wday}｜會議室 {rid}｜{time_str}｜用途：{purpose}"
            var = tk.BooleanVar()
            self.records.append((booking_id, wday, sid, rid, text))
            self.vars.append(var)

        # 排序（週一到週五）
        weekday_order = {"週一": 1, "週二": 2, "週三": 3, "週四": 4, "週五": 5}
        self.records.sort(key=lambda x: (weekday_order.get(x[1], 99), x[3], x[2]))

        if not self.records:
            tk.Label(self.list_frame, text="查無固定預約紀錄", font=("Arial", 12), bg="white", fg="gray").pack()
            return

        self.lbl_result_title.config(text=f"ID: {user_id} ，固定預約紀錄如下：")

        for i, (bid, _, _, _, text) in enumerate(self.records):
            frame = tk.Frame(self.list_frame, bg="white")
            frame.pack(anchor="w", fill="x", pady=4)

            cb = tk.Checkbutton(frame, text=text, variable=self.vars[i], font=("Arial", 11),
                                bg="white", anchor="w", justify="left")
            cb.pack(anchor="w")

    def cancel_selected(self):
        selected_ids = [self.records[i][0] for i, var in enumerate(self.vars) if var.get()]
        if not selected_ids:
            messagebox.showwarning("提醒", "請勾選要取消的固定預約")
            return

        try:
            wb = load_workbook(FILENAME)
            ws = wb["FixedBooking"]
            updated = False

            for row in ws.iter_rows(min_row=2):
                if row[0].value in selected_ids:
                    if len(row) < 7:
                        ws.cell(row=row[0].row, column=7, value=True)
                    else:
                        row[6].value = True
                    updated = True

            if updated:
                # 初學者註解：取消固定預約 → 安全存檔
                if not safe_save(wb, FILENAME):
                    messagebox.showerror("錯誤", "儲存 Excel 檔案失敗，請先關閉 Excel 或檢查檔案權限。")
                    return
                messagebox.showinfo("成功", "選取的固定預約已成功取消")
                self.search()
            else:
                messagebox.showinfo("提示", "找不到要取消的紀錄，可能已被取消")

        except Exception as e:
            messagebox.showerror("錯誤", f"取消時發生錯誤：\n{e}")




# 將目前的 app_state 打包成乾淨 dict
def get_booking_data():
    return {
        "date": app_state["selected_date"],
        "time_slots": app_state["selected_slots"],
        "room_id": app_state["selected_room"],
        "user_id": app_state["user_id"],
        "purpose": app_state["purpose"]
    }


# 清除 TempLock 中過期的鎖定資料
def cleanup_expired_locks(filename=FILENAME):
    wb = load_workbook(filename)

    # ✅ 如果沒有 TempLock 工作表，就自動建立並跳過本次清除
    if "TempLock" not in wb.sheetnames:
        ws = wb.create_sheet("TempLock")
        ws.append(["UserID", "Date", "SlotID", "RoomID", "Status", "Timestamp"])  # 加上欄位列
        safe_save(wb, FILENAME)
        print("TempLock 表不存在，自動建立完成，跳過本次清除")
        return  # 跳出函式，避免處理空資料

    # 若已存在就繼續正常清除過期鎖定
    ws = wb["TempLock"]
    now = datetime.now()
    rows_to_keep = [ws[1]]  # 保留欄位列（標題）

    for row in ws.iter_rows(min_row=2):
        timestamp_str = row[5].value
        try:
            lock_time = datetime.strptime(timestamp_str, "%Y/%m/%d %H:%M:%S")
            if (now - lock_time).total_seconds() < LOCK_EXPIRY_SECONDS:
                rows_to_keep.append(row)
        except:
            pass  # 無效或壞資料就略過

    # 清除後寫回有效資料
    ws.delete_rows(2, ws.max_row)
    for row in rows_to_keep[1:]:
        ws.append([cell.value for cell in row])

    safe_save(wb, filename)


# 將暫時鎖定資料寫入 TempLock 表
def lock_room(token, date, slot_ids, room_id, filename=FILENAME):
    wb = load_workbook(filename)
    ws = wb["TempLock"]

    timestamp_str = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    for slot_id in slot_ids:
        ws.append([token, date, slot_id, room_id, "LOCKING", timestamp_str])
    # 初學者註解：建立暫時鎖也用安全存檔，避免半寫入
    if not safe_save(wb, filename):
        messagebox.showerror("錯誤", "建立暫時鎖定失敗，請稍後再試。")
        return

# 釋放某使用者的所有鎖定資料
def release_token_locks(token, filename=FILENAME):
    wb = load_workbook(filename)
    ws = wb["TempLock"]

    rows_to_keep = [ws[1]]  # 保留欄位列（第1列標題）

    for row in ws.iter_rows(min_row=2):
        if row[0].value != token:
            rows_to_keep.append(row)

    # 清空後重建
    ws.delete_rows(2, ws.max_row)
    for row in rows_to_keep[1:]:
        ws.append([cell.value for cell in row])

    safe_save(wb, filename)

# ✅ 啟動時初始化 Excel
init_excel_file()
def has_real_data():
    """確認會議室與時段中至少有一筆不是『範例』的正式資料"""
    try:
        wb = load_workbook(FILENAME)

        # 檢查 MeetingRooms 資料列是否非空且用途不是範例
        ws_rooms = wb["MeetingRooms"]
        real_rooms = [
            row for row in ws_rooms.iter_rows(min_row=2, values_only=True)
            if row and "範例" not in str(row[5])  # 用途欄
        ]

        # 檢查 TimeSlots 資料列是否非空且備註不是範例
        ws_slots = wb["TimeSlots"]
        real_slots = [
            row for row in ws_slots.iter_rows(min_row=2, values_only=True)
            if len(row) >= 4 and "範例" not in str(row[3])  # 備註欄
        ]

        return len(real_rooms) > 0 and len(real_slots) > 0

    except Exception as e:
        print(f"資料檢查錯誤：{e}")
        return False


# 執行主程式
if is_excel_file_locked(FILENAME):
    messagebox.showerror("檔案鎖定", "meeting_schedule.xlsx 正在被 Excel 或其他程式使用，請先關閉檔案再執行。")
    exit()
def show_large_error(title, message):
    root = tk.Tk()
    root.title(title)
    root.geometry("500x400")
    root.configure(bg="white")
    root.resizable(False, False)

    tk.Label(
        root,
        text=title,
        font=("Arial", 18, "bold"),
        fg="red",
        bg="white"
    ).pack(pady=(20, 10))

    text_frame = tk.Frame(root, bg="white")
    text_frame.pack(fill="both", expand=True, padx=20, pady=10)

    text_box = tk.Text(
        text_frame,
        font=("Arial", 12),
        wrap="word",
        bg="white",
        relief="flat",
        borderwidth=0
    )
    text_box.insert("1.0", message)
    text_box.config(state="disabled")
    text_box.pack(fill="both", expand=True)

    tk.Button(
        root,
        text="關閉程式",
        command=root.destroy,
        font=("Arial", 12),
        bg="#ef4444",
        fg="white",
        relief="flat",
        width=15
    ).pack(pady=(10, 30))

    root.mainloop()


if __name__ == "__main__":
    init_excel_file()

    if not has_real_data():
        show_large_error(
            "尚未設定正式資料",
            "系統偵測到目前 Excel 檔案內仍是範例資料，請依以下說明操作後再重新啟動：\n\n"
            "【請操作以下兩項】\n"
            "1. 打開 meeting_schedule.xlsx\n"
            "2. 在『MeetingRooms』工作表中：\n"
            "   - 刪除或修改「用途」欄為『範例』的那一列\n"
            "3. 在『TimeSlots』工作表中：\n"
            "   - 刪除或修改「備註」欄為『範例』的那一列\n\n"
            "請至少各建立一筆正式資料後，再重新啟動程式。"
        )
        exit()

    app = MeetingApp()
    app.mainloop()
