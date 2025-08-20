from datetime import datetime, timedelta
from openpyxl import load_workbook
import tkinter as tk
from excel_manager import safe_save  # 初學者註解：安全寫入，先寫暫存檔→原子覆蓋→產生 .bak


FILENAME = "meeting_schedule.xlsx"
LOCK_EXPIRY_SECONDS = 20



# 載入時段
def load_time_slots(filename=FILENAME):
    wb = load_workbook(filename)
    ws = wb["TimeSlots"]
    time_slots = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        slot_id, time_range, disabled, *_ = row
        if str(disabled).strip().upper() != "TRUE":
            time_slots[int(slot_id)] = time_range
    return time_slots

# 清除過期鎖定
def cleanup_expired_locks(filename=FILENAME):
    wb = load_workbook(filename)

    if "TempLock" not in wb.sheetnames:
        ws = wb.create_sheet("TempLock")
        ws.append(["UserID", "Date", "SlotID", "RoomID", "Status", "Timestamp"])
        safe_save(wb, filename)
        print("TempLock 表不存在，自動建立完成，跳過本次清除")
        return

    ws = wb["TempLock"]
    now = datetime.now()
    rows_to_keep = [ws[1]]

    for row in ws.iter_rows(min_row=2):
        timestamp_str = row[5].value
        try:
            lock_time = datetime.strptime(timestamp_str, "%Y/%m/%d %H:%M:%S")
            if (now - lock_time).total_seconds() < LOCK_EXPIRY_SECONDS:
                rows_to_keep.append(row)
        except:
            pass

    ws.delete_rows(2, ws.max_row)
    for row in rows_to_keep[1:]:
        ws.append([cell.value for cell in row])

    safe_save(wb, filename)
class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None

    def show(self):
        if self.tipwindow:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + 20

        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)        # 無邊框視窗
        tw.wm_geometry(f"+{x}+{y}")

        label = tk.Label(
            tw,
            text=self.text,
            justify='left',
            background="#f9fafb",           # ✅ 表格同色系灰白底
            foreground="#111827",           # ✅ 深灰主字體
            relief="solid",
            borderwidth=1,
            highlightbackground="#d1d5db",  # ✅ 與表格線條一致
            highlightthickness=1,
            font=("Arial", 10),
            padx=10,
            pady=6
        )
        label.pack()

    def hide(self):
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None


def render_schedule_table(frame, time_slots, dates, bookings_by_day_slot, all_room_ids):
    """
    將預約狀況渲染至傳入的 tkinter Frame
    - frame: 要繪製到的 Frame
    - time_slots: dict, {slot_id: 時段文字}
    - dates: 一週內的五天日期字串 ["2025-07-29", ...]
    - bookings_by_day_slot: dict, {(date, slot_id): [room_id, ...]}
    - all_room_ids: 所有啟用中的會議室 ID 清單
    """
    slot_ids = sorted(time_slots.keys())

    # === 表頭列（時段 + 星期）===
    weekday_names = ["週一", "週二", "週三", "週四", "週五"]
    headers = ["時段"] + [
        f"{weekday_names[i]} ({datetime.strptime(date, '%Y/%m/%d').strftime('%m/%d')})"
        for i, date in enumerate(dates)
    ]

    for col, text in enumerate(headers):
        tk.Label(
            frame,
            text=text,
            font=("Arial", 11, "bold"),
            bg="white",
            fg="#111827",
            padx=10,
            pady=6,
            relief="flat"
        ).grid(row=0, column=col, sticky="nsew", padx=4, pady=4)

    # === 時段與每日格子 ===
    for r, sid in enumerate(slot_ids, start=1):
        # 時段名稱
        tk.Label(
            frame,
            text=time_slots[sid],
            font=("Arial", 10),
            bg="#f9fafb",
            fg="#111827",
            padx=8,
            pady=4
        ).grid(row=r, column=0, sticky="nsew", padx=4, pady=4)

        for c, date in enumerate(dates):
            used_rooms = bookings_by_day_slot.get((date, sid), [])
            if not used_rooms:
                color = "#AED581"  # 綠：全部可借
                border = "#9ccc65"
            elif len(set(used_rooms)) >= len(all_room_ids):
                color = "#E57373"  # 紅：無可借
                border = "#ef5350"
            else:
                color = "#FFD54F"  # 黃：部分可借
                border = "#ffca28"

            tk.Label(
                frame,
                text="",
                bg=color,
                borderwidth=1,
                highlightbackground=border,
                relief="flat",
                padx=6,
                pady=6
            ).grid(row=r, column=c + 1, sticky="nsew", padx=4, pady=4)

    # 平均配置欄寬與列高
    for col in range(6):
        frame.grid_columnconfigure(col, weight=1)
    for row in range(len(slot_ids) + 1):
        frame.grid_rowconfigure(row, weight=1)

def render_boss_table(frame):
    from datetime import datetime, timedelta
    from openpyxl import load_workbook

    # 清空舊畫面元件
    for widget in frame.winfo_children():
        widget.destroy()

    wb = load_workbook(FILENAME)
    time_slots = load_time_slots()
    slot_ids = sorted(time_slots.keys())

    # 取得本週一 ~ 週五的日期
    today = datetime.today()
    monday = today - timedelta(days=today.weekday())
    dates = [(monday + timedelta(days=i)).strftime("%Y/%m/%d") for i in range(5)]

    # 產生日期對應的中文星期
    weekday_names = ["週一", "週二", "週三", "週四", "週五"]
    date_to_wday = {
        date: weekday_names[datetime.strptime(date, "%Y/%m/%d").weekday()]
        for date in dates
    }

    # === 新增：取得「不可外借」會議室的 ID 清單 ===
    non_external_rooms = set()
    ws_rooms = wb["MeetingRooms"]
    for row in ws_rooms.iter_rows(min_row=2, values_only=True):
        room_id, *_ , allow_external = row
        if str(allow_external).strip().upper() != "TRUE":
            non_external_rooms.add(room_id)

    # === 取得固定預約資料 ===
    fixed_map = {}
    if "FixedBooking" in wb.sheetnames:
        ws = wb["FixedBooking"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 6:
                continue
            if len(row) >= 7 and row[6] is True:
                continue  # ❌ 已取消固定預約，略過

            booking_id, wday, sid, rid, uid, purpose = row[:6]

            try:
                sid = int(sid)
            except:
                continue

            # ✅ 加入過濾條件：只保留「不可外借」的會議室
            if isinstance(rid, str) and isinstance(wday, str) and rid in non_external_rooms:
                fixed_map.setdefault((wday.strip(), sid), []).append((rid, uid, purpose))

    # === 若無任何資料，顯示提示文字 ===
    if not fixed_map:
        tk.Label(
            frame,
            text="目前尚無任何不可外借的固定預約",
            font=("Arial", 14),
            bg="white",
            fg="gray"
        ).pack(pady=40)
        return

    # === 建立表頭列 ===
    headers = ["時段"] + [
        f"{date_to_wday[date]} ({datetime.strptime(date, '%Y/%m/%d').strftime('%m/%d')})"
        for date in dates
    ]
    for col, text in enumerate(headers):
        tk.Label(
            frame,
            text=text,
            font=("Arial", 11, "bold"),
            bg="white",
            fg="#111827",
            padx=10,
            pady=6
        ).grid(row=0, column=col, sticky="nsew", padx=4, pady=4)

    # === 建立格子內容 ===
    for r, sid in enumerate(slot_ids, start=1):
        bg_color = "#f9fafb" if r % 2 != 1 else "#e5e7eb"

        # 左側時段欄
        tk.Label(
            frame,
            text=time_slots[sid],
            font=("Arial", 10),
            bg=bg_color,
            fg="#111827",
            padx=8,
            pady=4
        ).grid(row=r, column=0, sticky="nsew")

        # 右側 5 天格子
        for c, date in enumerate(dates):
            wday = date_to_wday[date]
            cell_key = (wday, sid)
            has_booking = cell_key in fixed_map

            symbol = "●" if has_booking else ""

            # 預設紅色
            color = "#ef4444" if has_booking else "white"

            # ✅ 若用途中包含 MIS，改為主題藍色
            if has_booking:
                bookings = fixed_map[cell_key]
                if any("MIS" in str(purpose).upper() for _, _, purpose in bookings):
                    color = "#3B82F6"  # 主題藍


            # Tooltip 詳細內容
            details = []
            if has_booking:
                bookings = fixed_map[cell_key]
                for rid, uid, purpose in bookings:
                    # ✅ 若用途中有 MIS，加上提示
                    tag = ""
                    if "MIS" in str(purpose).upper():
                        tag = "\n⚠️ 此固定預約涉及 MIS 支援"

                    details.append(f"[固定] 會議室：{rid}\n預約人：{uid}\n用途：{purpose}{tag}")
            tooltip_text = "\n-------------------\n".join(details) if details else "尚未有預約"

            label = tk.Label(
                frame,
                text=symbol,
                fg=color,
                bg=bg_color,
                font=("Arial", 20, "bold")
            )
            label.grid(row=r, column=c + 1, sticky="nsew")

            if has_booking:
                tooltip = Tooltip(label, tooltip_text)
                label.bind("<Enter>", lambda e, tip=tooltip: tip.show())
                label.bind("<Leave>", lambda e, tip=tooltip: tip.hide())

    # 統一欄寬與列高
    for col in range(6):
        frame.grid_columnconfigure(col, weight=1)
    for row in range(len(slot_ids) + 1):
        frame.grid_rowconfigure(row, weight=1)




def render_weekly_table(frame):
    cleanup_expired_locks()

    # 清空原有內容
    for widget in frame.winfo_children():
        widget.destroy()

    time_slots = load_time_slots()
    slot_ids = sorted(time_slots.keys())

    today = datetime.today()
    monday = today - timedelta(days=today.weekday())
    dates = [(monday + timedelta(days=i)).strftime("%Y/%m/%d") for i in range(5)]

    wb = load_workbook(FILENAME)
    ws_schedule = wb["Schedule"]
    ws_rooms = wb["MeetingRooms"]

    all_room_ids = [
        row[0] for row in ws_rooms.iter_rows(min_row=2, values_only=True)
        if str(row[6]).strip().upper() != "TRUE"  # ✅ 未停用
        and str(row[7]).strip().upper() == "TRUE"  # ✅ 可外借
    ]


    room_name_map = {
        row[0]: row[1] for row in ws_rooms.iter_rows(min_row=2, values_only=True)
    }
            # 取得固定預約資料
    fixed_bookings = []
    if "FixedBooking" in wb.sheetnames:
        ws_fixed = wb["FixedBooking"]
        for row in ws_fixed.iter_rows(min_row=2, values_only=True):
            fixed_bookings.append(row)  # [BookingID, Weekday, SlotID, RoomID, UserID, Purpose]

    bookings_by_day_slot = {}
    valid_records_exist = False  # ✅ 是否有有效預約資料

    for row in ws_schedule.iter_rows(min_row=2, values_only=True):
        date, slot_str, room_id, canceled = row[1], row[2], row[3], row[6]
        if canceled or not slot_str:
            continue
        try:
            for sid in map(int, str(slot_str).split(',')):
                bookings_by_day_slot.setdefault((date, sid), []).append(room_id)
                valid_records_exist = True
        except:
            continue
            # 🔽 加入固定預約的房間（轉換為對應日期）
        ws_fixed = wb["FixedBooking"]
        for row in ws_fixed.iter_rows(min_row=2, values_only=True):
            if len(row) >= 7 and row[6] is True:
                continue  # ✅ 已取消，跳過
            _, wday, sid, rid, *_ = row
            if not isinstance(sid, int) or not rid or rid not in all_room_ids:
                continue
            # 轉換星期幾為對應的日期
            for i, weekday_str in enumerate(["週一", "週二", "週三", "週四", "週五"]):
                if wday == weekday_str:
                    date = dates[i]
                    bookings_by_day_slot.setdefault((date, sid), []).append(rid)
                    break  # 找到就不用再比對下去


    # ✅ 若無預約資料，顯示提示文字並離開
    if not valid_records_exist:
        tk.Label(
            frame,
            text="目前尚無任何預約紀錄",
            font=("Arial", 14),
            bg="white",
            fg="gray"
        ).pack(pady=40)
        return

    # === 表頭列 ===
    weekday_names = ["週一", "週二", "週三", "週四", "週五"]
    headers = ["時段"] + [
        f"{weekday_names[i]} ({datetime.strptime(dates[i], '%Y/%m/%d').strftime('%m/%d')})"
        for i in range(5)
    ]
    for col, text in enumerate(headers):
        tk.Label(
            frame,
            text=text,
            font=("Arial", 11, "bold"),
            bg="white",
            fg="#111827",
            padx=10,
            pady=6
        ).grid(row=0, column=col, sticky="nsew", padx=4, pady=4)

    # === 表格內容 ===
    for r, sid in enumerate(slot_ids, start=1):
        bg_color = "#f9fafb" if r % 2 != 1 else "#e5e7eb"
        tk.Label(
            frame,
            text=time_slots[sid],
            font=("Arial", 10),
            bg=bg_color,
            fg="#111827",
            padx=8,
            pady=4
        ).grid(row=r, column=0, sticky="nsew")

        for c, date in enumerate(dates):
            used_rooms = bookings_by_day_slot.get((date, sid), [])
            total_rooms = len(all_room_ids)
            used_count = len(set(used_rooms))

            if used_count == 0:
                symbol = "●"
                color = "#4caf50"  # 綠
            elif used_count >= total_rooms:
                symbol = "●"
                color = "#ef4444"  # 紅
            else:
                symbol = "●"
                color = "#f59e0b"  # 黃

            slot_name = time_slots[sid]
            details = []

            # 加入「一般預約」資訊
            for row in ws_schedule.iter_rows(min_row=2, values_only=True):
                b_date, b_slots, b_room, b_user, b_purpose, b_cancel = row[1], row[2], row[3], row[4], row[5], row[6]
                if b_cancel or b_date != date or not b_slots:
                    continue
                if b_room not in all_room_ids:
                    continue  # ✅ 不在可外借房間清單，跳過
                try:
                    booked_sids = set(map(int, str(b_slots).split(',')))
                except:
                    continue
                if sid in booked_sids:
                    room_name = room_name_map.get(b_room, "")
                    details.append(f"[一般] 會議室：{b_room}（{room_name}）\n預約人：{b_user}\n用途：{b_purpose}")

            # 加入「固定預約」資訊
            weekday_str = ["週一", "週二", "週三", "週四", "週五"][c]  # 第 c 欄對應的星期幾
            for fb in fixed_bookings:
                if len(fb) >= 7 and fb[6] is True:
                    continue  # ✅ 已取消，跳過

                wday, fsid, froom, fuser, fpurpose = fb[1:6]

                if wday == weekday_str and fsid == sid:
                    if froom not in all_room_ids:
                        continue  # ✅ 不可外借會議室，略過
                    room_name = room_name_map.get(froom, "")
                    
                        # ✅ 若用途中包含 MIS，加入提示
                    tag = ""
                    if "MIS" in str(fpurpose).upper():
                        tag = "\n⚠️ 此固定預約涉及 MIS 支援"

                    details.append(
                        f"[固定] 會議室：{froom}（{room_name}）\n預約人：{fuser}\n用途：{fpurpose}{tag}"
                    )



            tooltip_text = "\n-------------------\n".join(details) if details else "尚未有預約"
            bg_color = "#f9fafb" if r % 2 != 1 else "#e5e7eb"  # ✅ 交錯灰底
            label = tk.Label(
                frame,
                text=symbol,
                fg=color,
                bg=bg_color,
                font=("Arial", 20, "bold")
            )
            label.grid(row=r, column=c + 1, sticky="nsew")

            # ✅ 若為紅或黃，才顯示 Tooltip
            if color in ("#ef4444", "#f59e0b"):  # 紅 or 黃
                tooltip = Tooltip(label, tooltip_text)
                label.bind("<Enter>", lambda e, tip=tooltip: tip.show())
                label.bind("<Leave>", lambda e, tip=tooltip: tip.hide())

    for col in range(6):
        frame.grid_columnconfigure(col, weight=1)
    for row in range(len(slot_ids) + 1):
        frame.grid_rowconfigure(row, weight=1)


